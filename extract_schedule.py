#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
加古川市ゴミ出し日程抽出モジュール
Excelファイルから鳩里１地区のゴミ出し日程を抽出してJSON形式で保存
"""

import openpyxl
import json
import calendar
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
import sys
import os

# UTF-8 encoding対応
if sys.stdout.encoding != 'utf-8':
    sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)


class GomiScheduleExtractor:
    """加古川市ゴミ出し日程抽出クラス"""

    # ゴミ種別マッピング（Excelカラム位置）
    GARBAGE_TYPES = {
        '燃やすごみ': 'burnable',  # 毎週月・木
        '燃やさないごみ': ('non_burnable', 3),  # Column C
        'かん': ('cans', 4),  # Column D
        'びん': ('bottles', 5),  # Column E
        '剪定枝・草': ('branches', 6),  # Column F
        'ペットボトル': ('pet_bottles', 8),  # Column H
        '紙類': ('paper', 9),  # Column I
        '蛍光灯・電池': ('fluorescent', 11),  # Column K
    }

    def __init__(self, excel_path: str, year: int = 2026):
        """
        初期化

        Args:
            excel_path: Excelファイルのパス
            year: 対象年度（デフォルト2026年）
        """
        self.excel_path = excel_path
        self.year = year
        self.schedule = {}

    def extract(self) -> Dict:
        """
        Excelファイルからスケジュール情報を抽出

        Returns:
            ゴミ出し日程を含む辞書
        """
        print(f"ファイルを開いています: {self.excel_path}")
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb['2']  # シート2（鳩里１）

        print("スケジュール情報を抽出中...\n")

        # 1. 燃やすごみの固定スケジュール（毎週月・木）
        self.schedule['燃やすごみ'] = self._get_burnable_schedule()
        print(f"✓ 燃やすごみ: 毎週月・木曜日")

        # 2. その他のゴミ種別の月別スケジュール
        for jp_name, (en_name, col) in [
            ('燃やさないごみ', ('non_burnable', 3)),
            ('かん', ('cans', 4)),
            ('びん', ('bottles', 5)),
            ('剪定枝・草', ('branches', 6)),
            ('ペットボトル', ('pet_bottles', 8)),
            ('紙類', ('paper', 9)),
            ('蛍光灯・電池', ('fluorescent', 11)),
        ]:
            dates = self._extract_monthly_dates(ws, col)
            self.schedule[jp_name] = dates
            print(f"✓ {jp_name}: {len(dates)}件の日程")

        return self.schedule

    def _get_burnable_schedule(self) -> List[str]:
        """
        燃やすごみのスケジュール（毎週月・木曜日）を生成

        Returns:
            年間の燃やすごみ出し日のリスト
        """
        # 毎週月・木曜日のパターン
        burnable_dates = []

        # 対象年の1月1日から12月31日まで
        start_date = datetime(self.year, 1, 1)
        end_date = datetime(self.year, 12, 31)

        current = start_date
        while current <= end_date:
            # 月曜日(weekday=0)または木曜日(weekday=3)
            if current.weekday() in [0, 3]:
                burnable_dates.append(current.strftime('%Y-%m-%d'))
            current += timedelta(days=1)

        return burnable_dates

    def _extract_monthly_dates(self, ws, col: int) -> Dict[int, List[int]]:
        """
        月別スケジュール（特定日付）を抽出

        Args:
            ws: ワークシート
            col: 対象カラム番号（1ベース）

        Returns:
            {月: [日付リスト]} の辞書
        """
        monthly_dates = {}

        # 行15-27が月別データ
        for row_idx in range(15, 28):
            month_cell = ws.cell(row=row_idx, column=1)
            date_cell = ws.cell(row=row_idx, column=col)

            if month_cell.value is not None:
                try:
                    month = int(month_cell.value)
                    if date_cell.value is not None:
                        day = int(date_cell.value)
                        if month not in monthly_dates:
                            monthly_dates[month] = []
                        monthly_dates[month].append(day)
                except (ValueError, TypeError):
                    pass

        return monthly_dates

    def get_garbage_for_date(self, date: datetime) -> List[str]:
        """
        指定された日付のゴミ出し種別を取得

        Args:
            date: 対象日付

        Returns:
            該当するゴミ種別のリスト
        """
        garbage_types = []

        # 燃やすごみをチェック
        if date.strftime('%Y-%m-%d') in self.schedule.get('燃やすごみ', []):
            garbage_types.append('燃やすごみ')

        # その他のゴミをチェック
        for gomi_name, monthly_data in self.schedule.items():
            if gomi_name == '燃やすごみ':
                continue

            if isinstance(monthly_data, dict):
                month_dates = monthly_data.get(date.month, [])
                if date.day in month_dates:
                    garbage_types.append(gomi_name)

        return garbage_types

    def save_to_json(self, output_path: str):
        """
        抽出したスケジュールをJSON形式で保存

        Args:
            output_path: 出力ファイルパス
        """
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # JSONシリアライズ用に整形
        schedule_for_json = {
            'year': self.year,
            'generated_at': datetime.now().isoformat(),
            'schedule': {}
        }

        for gomi_name, data in self.schedule.items():
            if isinstance(data, dict):
                # 月別スケジュール
                schedule_for_json['schedule'][gomi_name] = data
            else:
                # 日付リスト
                schedule_for_json['schedule'][gomi_name] = data

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(schedule_for_json, f, ensure_ascii=False, indent=2)

        print(f"\n✓ スケジュールを保存しました: {output_path}")


def main():
    """メイン処理"""

    # Excelファイルのパス（WebFetchでダウンロードされたもの）
    excel_path = r"C:\Users\user\.claude\projects\C--Users-user-Downloads\dc4c8676-f7b3-460b-8b37-d134a8eb3aeb\tool-results\webfetch-1778334654076-r5kv4t.xlsx"

    # 出力先
    output_dir = r"C:\Users\user\gomi_reminder_system"
    output_file = os.path.join(output_dir, "schedule_cache.json")

    try:
        # 抽出実行
        extractor = GomiScheduleExtractor(excel_path, year=2026)
        schedule = extractor.extract()

        # JSON保存
        extractor.save_to_json(output_file)

        # サマリー表示
        print("\n" + "="*50)
        print("スケジュール抽出完了")
        print("="*50)

        # テスト：明日のゴミを確認
        tomorrow = datetime.now() + timedelta(days=1)
        garbage_tomorrow = extractor.get_garbage_for_date(tomorrow)
        print(f"\n今日: {datetime.now().strftime('%Y年%m月%d日（%A）')}")
        print(f"明日: {tomorrow.strftime('%Y年%m月%d日（%A）')}")
        if garbage_tomorrow:
            print(f"明日のゴミ: {', '.join(garbage_tomorrow)}")
        else:
            print("明日はゴミ出し日ではありません")

        return True

    except FileNotFoundError:
        print(f"エラー: ファイルが見つかりません: {excel_path}")
        return False
    except Exception as e:
        print(f"エラー: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)
